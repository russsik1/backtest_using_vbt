import multiprocessing, itertools, concurrent.futures
from typing import List, Tuple, Optional, Dict
from enum import Enum
from pydantic import BaseModel, field_validator
import vectorbtpro as vbt
import numpy as np
import pandas as pd
from tqdm.notebook import tqdm
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


class Direction(str, Enum):
    LONG='LONG'
    SHORT='SHORT'




class Params(BaseModel):
    open: np.array
    high: np.array
    low: np.array
    close: np.array
    entries: np.array
    fees: float=.0004
    slippage: float=.0000
    init_cash: float=10000
    size: float=100
    direction: Direction=Direction.LONG
    open_date_time: Optional[np.array] = None

    class Config:
        arbitrary_types_allowed = True

    @field_validator('entries', mode='before')
    def validate_entries(cls, v):
        return v.astype(bool)


class GroupParams(Params):
    exit_groups: List[Tuple]


class CommonParams(Params):
    take_profits: np.array
    stop_losses: np.array
    trailing_stop_activation: np.array
    trailing_stops: np.array


    def get_exit_groups(self, group_size: int=10) -> List[GroupParams]:
        all_combinations = list(itertools.product(self.take_profits, self.stop_losses, self.trailing_stop_activation, self.trailing_stops)) # Генерация всех возможных комбинаций
        groups: List[GroupParams] = []
        for i in range(0, len(all_combinations), group_size):
            groups.append(GroupParams(
                open_date_time=self.open_date_time,
                open=self.open,
                high=self.high,
                low=self.low,
                close=self.close,
                entries=self.entries,
                fees=self.fees,
                slippage=self.slippage,
                init_cash=self.init_cash,
                size=self.size,
                direction=self.direction,
                exit_groups=all_combinations[i:i + group_size]
            ))        
        return groups






class Backtest:

    def __init__(self):
        pass

    
    @staticmethod
    def run_group(params: GroupParams) -> pd.DataFrame:
        stat_df = pd.DataFrame()
        trades_df = pd.DataFrame()
        for group in params.exit_groups:
            if pd.isnull(group[3]) and pd.isnull(group[2]):
                continue
            pf = vbt.Portfolio.from_signals(
                close=params.close, 
                open=params.open, 
                high=params.high, 
                low=params.low, 
                entries=params.entries if params.direction==Direction.LONG else None,
                short_entries=params.entries if params.direction==Direction.SHORT else None, 
                tp_stop=group[0],
                sl_stop=group[1],
                tsl_stop=group[2],
                tsl_th=group[3],
                init_cash=params.init_cash, 
                fees=params.fees, 
                slippage=params.slippage, 
                size=params.size,
                freq='1min'
            )
            tp_sl_th_ts = f'{group[0]} {group[1]} {group[2]} {group[3]}'.replace('None', '____').replace('nan', '____')
            stats = pf.stats()
            row_df = pd.concat([
                pd.DataFrame([[tp_sl_th_ts, *group]], columns=['tp_sl_th_ts', 'tp', 'sl', 'th', 'ts']),
                pd.DataFrame(stats).transpose().drop(columns=['Start', 'End']),
            ], axis=1)
            stat_df = pd.concat([stat_df, row_df])
            trades = pf.get_trades().records
            if len(trades)>0:
                trades['col'] = tp_sl_th_ts
                trades['entry_idx'] = pd.to_datetime(params.open_date_time/1000, unit='s', utc=True).astype(str)[trades['entry_idx'].values]
                trades['exit_idx'] = pd.to_datetime(params.open_date_time/1000, unit='s', utc=True).astype(str)[trades['exit_idx'].values]
                trades = trades.rename(columns={'col': 'tp_sl_th_ts', 'entry_idx': 'entry_date', 'exit_idx': 'exit_date'})
                trades = trades.drop(columns=['id', 'direction', 'status', 'parent_id'])
                trades_df = pd.concat([trades_df, trades])
        return stat_df, trades_df




    def run(self, common_params: CommonParams, cpu_usage_weight: float=.8, group_size: int=10) -> pd.DataFrame:
        args = [(group_params,) for group_params in common_params.get_exit_groups(group_size)]
        result_df_parts = []

        with concurrent.futures.ProcessPoolExecutor(max_workers=int(multiprocessing.cpu_count() * cpu_usage_weight)) as executor:
            futures = [executor.submit(self.run_group, *arg) for arg in args]
            for future in tqdm(concurrent.futures.as_completed(futures), total=len(futures)):
                result_df_parts.append(future.result())
        return {'stat': pd.concat([x[0] for x in result_df_parts]), 'trades': pd.concat([x[1] for x in result_df_parts])}


    def save(self, filename: str, dfs: Dict[str, pd.DataFrame]):
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        for sheetname, df in dfs.items():
            ws = wb.create_sheet(title=sheetname)
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
        wb.save(f'backtest_results/{filename}')